# coding=utf-8
import pandas as pd
import numpy as np
import time
from datetime import datetime, timedelta
import re as regex

from openpyxl import load_workbook


path = '/Users/panc/Desktop/ENCCO销售实战营——分组积分情况.xlsx'
data = pd.read_excel(path)

wb = load_workbook(filename = path)
sheet_names = wb.sheetnames
name = sheet_names[0]
sheet_ranges = wb[name]

data_origin = pd.DataFrame(sheet_ranges.values)#读取单元格源值（公式）


team_list = [] #团队列表
team_score = []#团队总分

crowd = [] # 个人排名

crowd_score_day = []#个人分数数据（按每天 单元格）

#遍历下标1列
for index,value in enumerate(data.iloc[:,1]):
  if value == '序号':
    team_list.append(data.iloc[:,1][index-1])
  
  if value == '合计':
    team_score.append(data.iloc[index:index+1,:].iloc[:,-1].values[0])

#遍历下标2列
flag = False
_n = 0
for index,value in enumerate(data.iloc[:,2].astype('str').replace('nan','')):
  if (not flag) and value == '群内昵称':
    flag = True
    _n += 1
    continue

  if not value:
    flag = False
  
  if flag :
    current_row = data.iloc[index:index+1,:]
    score = current_row.iloc[:,-1].values[0]
    crowd.append((value,score,team_list[_n-1]))
    
    d = data_origin.iloc[index+1:index+1+1,3:13].replace('=','',regex=True).replace(np.nan,'0').astype('str').values[0] # .dropna(axis=1)
    crowd_score_day.append((value,d,team_list[_n-1],  )) # 读取个人所有分数数据


# 查找所有成员的加分记录



team_list = [(v,team_score[i]) for i,v in enumerate(team_list)] #合并团队数据

team_list = sorted(team_list,key = lambda x:x[1],reverse=True )# 团队

crowd = sorted(crowd,key = lambda x:x[1],reverse=True )#个人

def get_person_top5(list):
  return """
个人前五总分排名

🌹🌹🌹🌹🌹🌹🌹🌹🌹🌹🌹
第一名  {0[0]} {0[1]}分
🌹🌹🌹🌹🌹🌹🌹🌹🌹🌹🌹
第二名  {1[0]}  {1[1]}分
🌹🌹🌹🌹🌹🌹🌹🌹🌹🌹🌹
第三名 {2[0]}  {2[1]}分
🌷🌷🌷🌷🌷🌷🌷🌷🌷🌷🌷
第四名 {3[0]}  {3[1]}分
🌷🌷🌷🌷🌷🌷🌷🌷🌷🌷🌷
第五名 {4[0]}  {4[1]}分
🌷🌷🌷🌷🌷🌷🌷🌷🌷🌷🌷

统计时间:{time}
  """.format(*crowd,time = time.strftime('%Y年%m月%d日 %H:%M'))

def get_team_top(team_list):# todo 
  """
团队总分排名
🌹🌹🌹🌹🌹🌹🌹🌹🌹
🌹第一名  超凡队  133564分
🌹第二名   称霸队  72401分
🌹第三名   必胜队  69751分
🌹🌹🌹🌹🌹🌹🌹🌹🌹
第四名   亮剑队  63811分
第五名   巨人队  58727分
第六名   奇迹队   56810分
第七名    铿锵玫瑰队 52722分
第八名    战狼队  41920分
第九名    神锋队  37308分
第十名    飞虎队   37242分
第十一名    狼牙队    30041分
第十二名    雄鹰队    23081分   
第十三名    冠军队   21838分
第十四名    野狼队   20730分
第十五名    超越队  20112分
第十六名    猛虎队   14354分

统计时间:2019年7月10日19:00
  """
  res = []
  for index,value in enumerate(team_list):
    res.append('%s第%d名\t%s\t%d分' % ('🌹' if index < 3 else '', index+1,value[0],value[1]))
  res.insert(len(res)+1 if len(res) <= 3 else 3,'🌹🌹🌹🌹🌹🌹🌹🌹🌹')
  res.insert(0,'团队总分排名\n🌹🌹🌹🌹🌹🌹🌹🌹🌹')
  res.insert(len(res)+1,'\n统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M'))
  return '\n'.join(res)


def volume_list(crowd_score_day,scores_map,days = 'all',return_='number'):
  """ 
  获取每个成员的有效成交次数

  Arguments:
      crowd_score_day 成员成交的分数数据
      scores 有效的成交分数
      days str|list类型 需要统计那一天 默认all取所有天数；days = [0,1] 统计第一二天的数据
      return_ str 'number'|'desc' 返回成交数据  次数或者描述
  
  Returns:
      
  """
  valid_scores = [i for i in scores_map] #
  scores_pattern = '(%s)'%('|'.join([str(i) for i in valid_scores]))
  result = []# 
  
  for (name,data,team_name) in crowd_score_day:
    num = 0 #成交次数
    desc = []# 成交描述
    for ii,day_data in enumerate(data):
      if days == 'all' or (isinstance(days,list) and ii in days) :
        day_i = ii+1
        findall = regex.findall(r'(?:\D|^)%s(?=\D|$)' % scores_pattern,day_data)
        desc += [scores_map[str(i)] for i in findall]
        num += len(findall)
        
    if return_ == 'number':
      result.append( (name,num,team_name,','.join(data)) )
    else:
      result.append( (name,desc,team_name,','.join(data)) )
   
  return result

def hangegg(valid_volume_list):
  """挂蛋列表（未成交名单）
  挂蛋名单
《光荣榜》❗❗❗❗❗❗
[炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹]
王者荣耀易燃易爆团队
胡贤 胡小辉 唐诚 李志路
[炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹]
英雄联盟团队
苏炳蔚 何璇 卢金阳 肖茹馨
[炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹]

统计时间： 2020-02-17 15:05
  """
  result = {}
  for (name,num,team_name,string) in valid_volume_list:
    name = name.split('+')
    name = name[1]
    if num == 0:
      if team_name in result:
        result[team_name].append(name)
      else:
        result[team_name] = [name]
  desc = '挂蛋名单\n《光荣榜》❗❗❗❗❗❗\n[炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹]\n'
  for i in result:
    item = result[i]
    desc += '%s\n' % (i)
    desc += '%s\n'% ' '.join(item)
    desc += '[炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹][炸弹]\n'
  desc += '统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')

  return desc

def poegg_list(valid_volume_list):
  """破蛋列表（已成交）
  """
  result = {}
  for (name,num,team_name,string) in valid_volume_list:
    name = name.split('+')
    name = name[1]
    
    if num > 0:
      if team_name in result:
        result[team_name].append(name)
      else:
        result[team_name] = [name]

  desc = '破蛋名单\n《光荣榜》❗❗❗❗❗❗\n🌹🌹🌹🌹🌹🌹🌹🌹🌹\n'
  for i in result:
    item = result[i]
    desc += '%s\n' % (i)
    desc += '%s\n'% ' '.join(item)
    desc += '🌹🌹🌹🌹🌹🌹🌹🌹🌹\n'
  desc += '统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')
  return desc

def team_volume_list(valid_volume_list):
  """团队成交量列表
  """
  team_volume = {}# 团队总成交量
  for (name,num,team_name,string) in valid_volume_list:
    name = name.split('+')
    name = name[1]
    
    if team_name in team_volume:
        team_volume[team_name] += num
    else:
        team_volume[team_name] = num
  res = [(name,team_volume[name]) for name in team_volume]
  res = sorted(res,key=lambda x:x[1],reverse=True)
  desc = '累计团队成交量排名\n\n'
  desc += '\n'.join(['%s: %d' % (name,num) for name,num in res])
  desc += '\n\n统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')
  return desc

def team_volume_today_list(valid_volume_list,today_index = 2):
  """
  今天的团队交易量排名
  
连单接龙🎉🎉🎉
冠军队   ：4单
火狼队   ：
圆梦队   ：1单
巅峰队   ：  1单
鹰王队   ：3单
飞鹰队   ：1单  
战狼队   ：1单   
创奇迹队：    
天使队   ：5单     
狂狼队   ：6单     
卓凡队   ： 1单

  Arguments:
      valid_volume_list {[type]} -- [description]
  """
  team_volume = {}# 团队总成交量
  for (name,num,team_name,string) in valid_volume_list:
    name = name.split('+')
    name = name[1]
    
    if team_name in team_volume:
        team_volume[team_name] += num
    else:
        team_volume[team_name] = num
  res = [(name,team_volume[name]) for name in team_volume]

  res = sorted(res,key=lambda x:x[1],reverse=True)
  desc = '今日连单接龙🎉🎉🎉\n\n'
  desc += '\n'.join(['%s\t: %d单' % (name,num) for name,num in res])
  desc += '\n\n统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')
  return desc




def person_volume_list(valid_volume_list):
  """个人成交量列表
  """
  person_volume = []# 个人总成交量
  valid_volume_list_top = sorted(valid_volume_list,key=lambda x: x[1],reverse = True)
  desc = '个人成交量排名\n\n'
  # name.split('+')[1]
  desc += '\n'.join(['%s: %d' % (name,num) for (name,num,team,string) in valid_volume_list_top])
  desc += '\n\n统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')
  return desc

def person_volume_desc_list(volume_list__):
  """个人销售产品描述
  
  Arguments:
      volume_list__ {[type]} -- [description]
  
  Returns:
      [type] -- [description]
  """
  volume_list__ = volume_list(crowd_score_day,scores_map = scores_map ,days = 'all',return_='desc')
  res = {}
  desc_info = '个人成交产品信息\n\n'

  for (name,desc_data,team,data) in volume_list__:
    set_= set(desc_data)
    desc = {i:desc_data.count(i) for i in set_}
    info = ['\t%s: %d\n' % (i,desc[i]) for i in desc]
    if info:
      desc_info += '%s\n' % name
      desc_info += ''.join(info)
    res[name] = desc


  dd = pd.DataFrame(res).stack().unstack(0).replace(np.nan,'')
  print(dd)
  # dd.to_excel('/Users/panc/Desktop/person_vol.xlsx')
  # desc += '\n'.join(['%s: %d' % (name,num) for (name,num,team,string) in valid_volume_list_top])
  # desc += '\n\n统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')
  return desc_info

  return res


def team_poegg(volume_list_):
  """团队成员是否全部有成交
  """
  team_list = list(set([team for (name,num,team,string) in volume_list_]))
  for (name,num,team,string) in volume_list_:
    if num == 0:
      if team in team_list:
        index = team_list.index(team)
        del team_list[index]
  desc = '全员破蛋团队\n\n'
  if team_list:
    desc += '\n'.join(team_list)
  else:
    desc += '无\n'
  desc += '\n统计时间:%s'%time.strftime('%Y年%m月%d日 %H:%M')
  return desc


  
# print(get_team_top(team_list))
# print(crowd)

first_day = (2020, 2, 16) #指定活动第一天的日期  

# 有效成交积分和描述
scores_map = {
'65' :'一个体验装',
# '110' :'两个体验装', #有效的成交分数 110为两个一起成交的65 
'150':'总代|董事 定金1500',
'200':'董事定金2028',
'240':'总代2379',
'330':'总代2884',
'380':'总代3752',
'1150':'大区9800',
'2300':'董事19800',
'1390':'董事13848',
}

volume_list_ = volume_list(crowd_score_day,scores_map = scores_map ,days = 'all')

dif = datetime.now() - datetime(*first_day) #今天和first_day相比相差几天
today_index = dif.days
days_num = min(10,today_index + 1)# excel表格记录的当前
print('今天 %s 是比赛活动开始的第%d天' % (str(datetime.now()).split('.')[0],days_num) )
# print(today_index)
# today_index = 2 # 02.18
# today_index = 3 # 02.19
volume_list_today = volume_list(crowd_score_day,scores_map = scores_map ,days = [today_index]) # 统计今天的得分情况

person_top5_info = get_person_top5(crowd)
team_top_info = get_team_top(team_list)
hangegg_info = hangegg(volume_list_)# 挂蛋信息
poegg_info = poegg_list(volume_list_) #破蛋信息排名
team_volume_info = team_volume_list(volume_list_)# 团队成交量排名
person_volume_info = person_volume_list(volume_list_)#个人成交量排名
person_volume_desc_info = person_volume_desc_list(volume_list(crowd_score_day,scores_map = scores_map,return_='desc'))# 个人成交的产品描述
team_poegg_info = team_poegg(volume_list_) #全员成交的团队
team_volume_today_info = team_volume_today_list(volume_list_today)#今天累计成交



import pyperclip

info = '\n\n\n'.join([
# person_top5_info,
# team_top_info,
# person_volume_info,
# team_volume_info,
# team_volume_today_info , 
# team_poegg_info , 
# hangegg_info,
person_volume_desc_info
]) # 

# pyperclip.copy(info) #复制到粘贴板


print(  info )
